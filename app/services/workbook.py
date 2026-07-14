from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook, Workbook

from app.models import TtsRequest


REQUIRED_COLUMNS = {"text", "voice_id"}
OPTIONAL_COLUMNS = {
    "voice_name",
    "accent",
    "speech_context",
    "target_seconds",
    "wpm",
    "export_m4a",
    "enhance_text",
}


class WorkbookError(ValueError):
    pass


class WorkbookService:
    def parse_requests(self, path: Path) -> list[TtsRequest]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if "tts_requests" not in workbook.sheetnames:
            raise WorkbookError("Workbook must contain a sheet named 'tts_requests'.")
        sheet = workbook["tts_requests"]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise WorkbookError("Workbook is empty.")

        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        header_map = {header: index for index, header in enumerate(headers) if header}
        missing = REQUIRED_COLUMNS - set(header_map)
        if missing:
            raise WorkbookError(f"Missing required column(s): {', '.join(sorted(missing))}.")

        requests: list[TtsRequest] = []
        for row_number, row in enumerate(rows[1:], start=2):
            values = {header: self._cell(row, index) for header, index in header_map.items()}
            if not any(values.values()):
                continue
            try:
                requests.append(
                    TtsRequest(
                        text=str(values.get("text") or ""),
                        voice_id=str(values.get("voice_id") or ""),
                        voice_name=self._optional_string(values.get("voice_name")),
                        accent=self._optional_string(values.get("accent")) or "neutral",
                        speech_context=self._optional_string(values.get("speech_context"))
                        or "outreach_conversational",
                        target_seconds=int(values.get("target_seconds") or 55),
                        wpm=int(values.get("wpm") or 135),
                        export_m4a=self._bool(values.get("export_m4a"), default=True),
                        enhance_text=self._bool(values.get("enhance_text"), default=False),
                    )
                )
            except Exception as exc:
                raise WorkbookError(f"Row {row_number} is invalid: {exc}") from exc
        return requests

    def write_results(self, path: Path, results: list[dict[str, Any]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "tts_results"
        headers = [
            "status",
            "text",
            "voice_id",
            "voice_name",
            "estimated_seconds",
            "actual_seconds",
            "warning_level",
            "warning_code",
            "mp3_url",
            "m4a_url",
            "error",
        ]
        sheet.append(headers)
        for result in results:
            warning = result.get("warning") or {}
            sheet.append(
                [
                    result.get("status"),
                    result.get("text"),
                    result.get("voice_id"),
                    result.get("voice_name"),
                    result.get("estimated_seconds"),
                    result.get("actual_seconds"),
                    warning.get("level"),
                    warning.get("code"),
                    result.get("mp3_url"),
                    result.get("m4a_url"),
                    result.get("error"),
                ]
            )
        workbook.save(path)

    def _cell(self, row: tuple, index: int):
        return row[index] if index < len(row) else None

    def _optional_string(self, value) -> str | None:
        if value is None:
            return None
        text = str(value).strip()
        return text or None

    def _bool(self, value, *, default: bool = False) -> bool:
        if isinstance(value, bool):
            return value
        if value is None:
            return default
        return str(value).strip().lower() in {"1", "true", "yes", "y", "m4a"}
